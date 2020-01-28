
# -*- coding: utf-8 -*-

import openpyxl
import csv
import re
import glob
import os
from openpyxl import Workbook

files = glob.glob(os.getcwd() + "/source/*.xlsx")
# print(files)
for fl in files:
    bookname = fl.split("/")[-1].split('.')[0]
    print(bookname)
    wb_obj = openpyxl.load_workbook(fl)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    
    # # Creating folder 
    # os.makedirs(bookname)
    # s_path = glob.glob(os.getcwd()+'/'+bookname +'/')
    
    list1 = []
    list2 = []
    for i in range(2, max_row + 1):
        book = sheet_obj.cell(row=i, column=1).value
        chapter = sheet_obj.cell(row=i, column=2).value
        token = sheet_obj.cell(row=i,column=3).value
        # print(chapter)
        if chapter == 'GEN' or chapter == 'EXO':
            list1.append(token)
        elif chapter == 'LEV' or chapter == 'NUM' or chapter == 'DEU':
            list2.append(token)
        else:
            pass
    # print(len(list1))
    # print(len(list2))
    
    
    unique_tokens = set(list2) - set(list1)
    print(len(unique_tokens))
    # open_book = Workbook()
    # sheet = open_book.active
    # for k in unique_tokens:
    #     sheet.append([k])
    # open_book.save('target.xlsx')
