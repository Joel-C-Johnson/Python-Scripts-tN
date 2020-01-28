# -*- coding: utf-8 -*-
import openpyxl
import csv
import re
import glob
import os

# Tsv files
Source_files = glob.glob(os.getcwd() + "/en_tn-v15/*.tsv")

for s_file in Source_files:
    bookName = s_file.split("/")[-1].split(".")[0]
    fileName = s_file.split("/")[-1]
    s_Path = glob.glob(os.getcwd() + "/en_tn-v15/" + str(fileName))
    t_Path = glob.glob(os.getcwd() + "/tN_Marathi/" + str(bookName) + ".xlsx")
    print(s_Path)
    print(t_Path)
    print(bookName)
    # Open Target File
    wb_obj = openpyxl.load_workbook(t_Path[0])
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    eng_row_count = 0
    hindi_row_count = 1


    label_file = (bookName + ".tsv")
    with open(label_file, 'w', encoding='utf-8') as tsv_file:
        twriter = csv.writer(tsv_file, delimiter='\t')
        with open(s_Path[0], 'r', encoding='utf-8') as tsvfile:
            reader = csv.reader(tsvfile, delimiter='\t')
            for rows in range(max_row):
                try:
                    row = next(reader)
                except:
                    break
                if row[8] == '':
                    pass
                else:
                    # dic1 = []
                    combine_cells = ''
                    eng_ocr_note = row[8]
                    split_eng_ocr = eng_ocr_note.split('\n')
                    find_array_len = len(split_eng_ocr)
                    if (len(split_eng_ocr) > 1):
                        subline_count = 0
                        array_count = find_array_len
                        for sub_lines in split_eng_ocr:
                            split_sublines = sub_lines.split('\t')
                            if sub_lines == '':
                                pass
                            elif subline_count == 0:
                                trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
                                hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
                                hin_ocr_note = str(hin_ocr_note0).strip()
                                combine_cells += hin_ocr_note + '\n'
                                subline_count += subline_count + 1
                                hindi_row_count += 1
                            elif array_count == 1:
                                hin_book = sheet_obj.cell(row=hindi_row_count, column=1).value
                                hin_book1 = hin_book.split("\n")[0]
                                hin_chaptr = sheet_obj.cell(row=hindi_row_count, column=2).value
                                hin_verse = sheet_obj.cell(row=hindi_row_count, column=3).value
                                trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
                                hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
                                hin_ocr_note = str(hin_ocr_note0).strip()
                                combine_cells += str(hin_book1).strip() + "\t" + str(hin_chaptr).strip() + "\t" + str(hin_verse).strip() + "\t" + str(split_sublines[3]) + str(split_sublines[4]) + "\t" + str(split_sublines[5]) + "\t" + str(split_sublines[6]) + "\t" + str(split_sublines[7]) + "\t" + str(hin_ocr_note)
                                hindi_row_count += 1
                            else:
                                hin_book = sheet_obj.cell(row=hindi_row_count, column=1).value
                                hin_book1 = hin_book.split("\n")[0]
                                hin_chaptr = sheet_obj.cell(row=hindi_row_count, column=2).value
                                hin_verse = sheet_obj.cell(row=hindi_row_count, column=3).value
                                trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
                                hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
                                hin_ocr_note = hin_ocr_note0.strip()
                                combine_cells += str(hin_book1).strip() + "\t" + str(hin_chaptr).strip() + "\t" + str(hin_verse).strip() + "\t" + str(split_sublines[3]) + str(split_sublines[4]) + "\t" + str(split_sublines[5]) + "\t" + str(split_sublines[6]) + "\t" + str(split_sublines[7]) + "\t" + str(hin_ocr_note) +'\n'
                                hindi_row_count += 1
                            array_count -= 1
                        find_link_source = re.findall(r'(\[\[\w+\:[\/\w+\-]*\]\])', eng_ocr_note)
                        find_link_target = re.findall(r'@', combine_cells)
                        edited_targetl = combine_cells
                        dic1 = []
                        if find_link_source:
                            for link in find_link_source:
                                if link == '':
                                    pass
                                else:
                                    dic1.append(link)
                            for k in dic1:
                                edited_targetl = edited_targetl.replace("@", k, 1)
                            rep_br = re.sub(r'\$', '<br>', edited_targetl)
                            twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()])        

                            print(eng_ocr_note)
                            print("------------------------------------------------------")
                            print(str(rep_br.strip()))
                            print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
                            print(" ")
                            print(" ")

                        else:
                            rep_br = re.sub(r'\$', '<br>', combine_cells)
                            twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()])

                            print(eng_ocr_note)
                            print("------------------------------------------------------")
                            print(str(rep_br.strip()))
                            print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
                            print(" ")
                            print(" ")

                    else:
                        find_link_source = re.findall(r'(\[\[\w+\:[\/\w+\-]*\]\])', eng_ocr_note)
                        trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
                        hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
                        hin_ocr_note = str(hin_ocr_note0).strip()
                        if hin_ocr_note == None:
                            break
                        find_link_target = re.findall(r'@',hin_ocr_note)
                        edited_targetl = hin_ocr_note 
                        hindi_row_count += 1
                        dic1 = []

                        if find_link_source:
                            for link in find_link_source:
                                if link == '':
                                    pass
                                else:
                                    dic1.append(link)   
                            for k in dic1:
                                edited_targetl = edited_targetl.replace("@", k, 1)
                            rep_br = re.sub(r'\$', '<br>', edited_targetl)
                            twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()]) 
                            print(eng_ocr_note)
                            print("------------------------------------------------------")
                            print(str(rep_br.strip()))
                            print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
                            print(" ")
                            print(" ")

                        else:
                            rep_br = re.sub(r'\$', '<br>', hin_ocr_note)
                            twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()])
                            print(eng_ocr_note)
                            print("------------------------------------------------------")
                            print(str(rep_br.strip()))
                            print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
                            print(" ")
                            print(" ")




