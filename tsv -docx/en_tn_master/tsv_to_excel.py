'''
The script is written for extracting the required data from .tsv files and create an excel files for the data.

	NOTE:
		Please Delete the Old md file before you run this code each time, 
		else the file will rewrite the same content again and again.
'''


import openpyxl
from openpyxl import load_workbook
import csv
import os
import re
import glob


'''Accessing the folder'''
os.chdir("tsv_files")

'''Fetching all .tsv files from the folder'''
files = glob.glob('*.tsv')
for file in files:
	rows = 1
	filename = re.sub(r"\.tsv","",file)
	'''Opening and reading the .tsv file'''
	with open(file,'r',encoding = 'utf-8') as tsvfile:
		reader = csv.reader(tsvfile, delimiter='\t')
		for row in reader:
			book = row[0]
			chapter = row[1]
			verse = row[2]
			hash_content = row[7]
			content1 = row[8]

			'''Using Regex for cleaning-up the unwanted contents'''
			edited = re.sub(r"<br>","$",content1)
			content = re.sub(r"(\[\[\w+\:[\/\w+\-]*\]\])","",edited)
			#hash_edit = re.sub(r"(#+)","\n\\1",edited_note)

			tsv_path = os.getcwd()
			os.chdir("..")

			if os.path.exists("excel_file"):
				os.chdir("excel_file")
				excel_path = os.getcwd()
			else:
				os.mkdir("excel_file")
				os.chdir("excel_file")
				excel_path = os.getcwd()

			excel_file = excel_path + "/" + filename + ".xlsx"

			if os.path.exists(excel_file):
				work_book = load_workbook(excel_file)
				worksheet = work_book.get_sheet_by_name('Sheet')
				work_book.active
				
				'''Writing the values row by row'''
				worksheet['A' + str(rows)] = book
				worksheet['B' + str(rows)] = chapter
				worksheet['C' + str(rows)] = verse
				#worksheet['D' + str(rows)] = hash_content
				worksheet['D' + str(rows)] = content
				work_book.save(excel_file)
				print(book,chapter,verse)
				rows += 1


			else:

				''' Create and loads the workbook'''
				work_book = openpyxl.Workbook()
				work_book.save(excel_file)
				work_book = load_workbook(excel_file)
				worksheet = work_book.get_sheet_by_name("Sheet")	
				
				'''Updating the sheet with column'''
				#worksheet['A1'] = 'Book'
				#worksheet['B1'] = 'Chapter'
				#worksheet['C1'] = 'Verse'
				#worksheet['D1'] = 'GL_Quote'
				#worksheet['E1'] = 'Notes'

				'''Writing the values row by row'''
				worksheet['A' + str(rows)] = book
				worksheet['B' + str(rows)] = chapter
				worksheet['C' + str(rows)] = verse
				#worksheet['D' + str(rows)] = hash_content
				worksheet['D' + str(rows)] = content
				work_book.save(excel_file)
				print(book,chapter,verse)
				rows += 1
				
			os.chdir(tsv_path)
